#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/1/13 10:49
# @Author  : Chocolate
# @Site    : 
# @File    : parse_json.py
# @Software: PyCharm

# python解析json，并把json数据写入excel

import os
import json
import xlwt
from openpyxl import Workbook

'''
xlrd和xlwt处理的是xls文件，单个sheet最大行数是65535，如果有更大需要的，建议使用openpyxl函数，最大行数达到1048576。
如果数据量超过65535就会遇到：ValueError: row index was 65536, not allowed by .xls format
'''
def get_json_info(path_name):
    book = xlwt.Workbook(encoding="utf-8", style_compression=0)  # 创建workbook对象，相当于一个文件
    sheet = book.add_sheet('图片分析', cell_overwrite_ok=True)  # 创建工作表
    col = ("关键字", "概率")
    row = 1
    for i in range(0,2):
        # 第0行，第i列
        sheet.write(0,i,col[i]) # 列名
    all_json = os.listdir(path_name)
    for json_name in all_json:
        # print(path_name + json_name)
        json_txt = open(path_name + json_name)
        json_info = json.load(json_txt)
        images = json_info['images'][0] # type:dict
        bi_concepts = images['bi-concepts'] # type:dict
        for i,(k,v) in enumerate(bi_concepts.items()):
            sheet.write(row,0,k)
            sheet.write(row,1,v)
            row += 1
            # 只取前五百个关键字
            if(i==499):
                break
    book.save(r"D:/parse_json.xls")

def get_json_file_openxl(path_name):
    # 创建一个excel文件
    wb = Workbook()
    # 创建一个表格
    sheet = wb.active
    col = ("关键字", "概率")
    row = 2
    for i in range(0, 2):
        # 第0行，第i列
        sheet.cell(1, i+1, col[i])  # 列名
    all_json = os.listdir(path_name)
    for json_name in all_json:
        # print(path_name + json_name)
        json_txt = open(path_name + json_name)
        json_info = json.load(json_txt)
        images = json_info['images'][0]  # type:dict
        bi_concepts = images['bi-concepts']  # type:dict
        for i, (k, v) in enumerate(bi_concepts.items()):
            if(row > 1048576):
                row = 1
                sheet2 = wb.create_sheet('sheet2')
                sheet = sheet2
            else:
                sheet.cell(row, 1, k)
                sheet.cell(row, 2, v)
                row += 1
                # 只取前五百个关键字
                if (i == 499):
                    break
    # 保存文件
    wb.save(r'D:/My_json_file.xlsx')

if __name__ == '__main__':
    # path_name = "D:/TourismData/政府层面/test_json/"
    # path_name = "D:/TourismData/政府层面/图片解析2.0/"
    # path_name = "D:/TourismData/旅游者层面/图片/图片解析/"
    path_name = 'D:/TourismData/json_file/'
    # path_name = 'D:/test_json/'
    # get_json_info(path_name)
    get_json_file_openxl(path_name)
    print("succeed")